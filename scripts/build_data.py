#!/usr/bin/env python3
"""
Baixa a planilha de Precos de Paridade de Importacao (PPI) da ANP,
normaliza as 4 abas semanais e gera os artefatos consumidos pelo dashboard.

Saidas (em docs/data/):
  ppi.json              -> dataset completo usado pelo dashboard
  ppi_<produto>.csv     -> serie longa por produto
  ppi_long.csv          -> serie longa consolidada dos 4 produtos
  meta.json             -> metadados da ultima execucao

Fonte: https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/
       precos/precos-de-paridade-de-importacao
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import sys
import unicodedata
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

XLSX_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/arq-ppi/ppi.xlsx"
)
PAGE_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/precos-de-paridade-de-importacao"
)

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "data"
RAW_DIR = ROOT / "data_raw"

DATE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")
HEADER_LABELS = {"data", "semana"}

# Aba na planilha -> (chave interna, rotulo curto)
SHEETS = [
    ("Gasolina R$ semanal", "gasolina", "Gasolina A"),
    ("Diesel R$ semanal", "diesel", "Diesel A S10"),
    ("QAV R$ semanal", "qav", "QAV"),
    ("GLP R$ kg semanal", "glp", "GLP"),
]

BR_TZ = timezone(timedelta(hours=-3))


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #
def norm(text) -> str:
    """Minusculo, sem acentos e sem espacos duplicados."""
    if text is None:
        return ""
    s = str(text).replace("\xa0", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()


def clean(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).replace("\xa0", " ")).strip()


def as_number(value):
    """Converte celula em float, tolerando texto com virgula decimal e '-'."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return None if f != f else f
    s = str(value).strip()
    if s in {"", "-", "--", "n/d", "N/D", "ND"}:
        return None
    s = s.replace("R$", "").replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        f = float(s)
    except ValueError:
        return None
    return None if f != f else f


def download(url: str) -> bytes:
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; anp-ppi-dashboard/1.0)",
        "Accept": "*/*",
    }
    last = None
    for attempt in range(4):
        try:
            r = requests.get(url, headers=headers, timeout=120)
            r.raise_for_status()
            if len(r.content) < 10000:
                raise RuntimeError(f"arquivo truncado: {len(r.content)} bytes")
            return r.content
        except Exception as exc:  # noqa: BLE001
            last = exc
            print(f"  tentativa {attempt + 1} falhou: {exc}", file=sys.stderr)
    raise SystemExit(f"nao foi possivel baixar {url}: {last}")


# --------------------------------------------------------------------------- #
# parsing
# --------------------------------------------------------------------------- #
def find_header(rows):
    """Linha de cabecalho ('Data' ou 'Semana') e indice da coluna de data."""
    for i, row in enumerate(rows[:15]):
        for j, cell in enumerate(row):
            if norm(cell) in HEADER_LABELS:
                return i, j
    raise ValueError("cabecalho ('Data'/'Semana') nao encontrado")


def read_locations(header_row, date_col: int):
    """Terminais = celulas nao vazias apos a data, ate a primeira lacuna.

    A ANP repete os mesmos terminais num segundo bloco (variacao % semanal),
    separado por uma coluna em branco; paramos na lacuna.
    """
    locations = []
    for j in range(date_col + 1, len(header_row)):
        name = clean(header_row[j])
        if not name:
            break
        locations.append({"col": j, "name": name})
    if not locations:
        raise ValueError("nenhum terminal encontrado no cabecalho")
    return locations


def parse_period(raw):
    """'05/11/2018 a 09/11/2018' -> (inicio ISO, fim ISO, rotulo)."""
    if isinstance(raw, datetime):
        iso = raw.date().isoformat()
        return iso, iso, raw.strftime("%d/%m/%Y")
    label = clean(raw)
    matches = DATE_RE.findall(label)
    if not matches:
        return None

    def iso(m):
        d, mth, y = m
        return f"{y}-{mth}-{d}"

    start, end = iso(matches[0]), iso(matches[-1])
    try:
        datetime.fromisoformat(start)
        datetime.fromisoformat(end)
    except ValueError:
        return None
    return start, end, label


def parse_sheet(ws, key: str, fallback_label: str) -> dict:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx, date_col = find_header(rows)
    locations = read_locations(rows[header_idx], date_col)

    def first_text(idx):
        if idx < 0 or idx >= len(rows):
            return ""
        for cell in rows[idx]:
            txt = clean(cell)
            if txt and norm(txt) != "variacao % em relacao a semana anterior":
                return txt
        return ""

    title = first_text(0) or fallback_label
    unit_raw = first_text(1)
    m = re.search(r"(R\$\s*/\s*[^\s,;]+)", unit_raw)
    unit = clean(m.group(1)) if m else (unit_raw or "R$")

    weeks, seen = [], set()
    for row in rows[header_idx + 1:]:
        if date_col >= len(row):
            continue
        period = parse_period(row[date_col])
        if not period:
            continue
        start, end, label = period
        if end in seen:
            continue
        values = []
        for loc in locations:
            v = as_number(row[loc["col"]]) if loc["col"] < len(row) else None
            values.append(round(v, 6) if v is not None else None)
        if all(v is None for v in values):
            continue
        seen.add(end)
        weeks.append({"start": start, "end": end, "label": label, "v": values})

    weeks.sort(key=lambda w: w["end"])
    if not weeks:
        raise ValueError(f"aba '{ws.title}' sem linhas de dados validas")

    return {
        "key": key,
        "label": fallback_label,
        "title": title,
        "unit": unit,
        "sheet": ws.title,
        "locations": [loc["name"] for loc in locations],
        "weeks": weeks,
    }


# --------------------------------------------------------------------------- #
# saidas
# --------------------------------------------------------------------------- #
def write_csvs(products: dict):
    consolidated = [
        ("produto", "unidade", "semana_inicio", "semana_fim", "terminal", "preco")
    ]
    for key, p in products.items():
        rows = [("semana_inicio", "semana_fim", "terminal", "preco")]
        for w in p["weeks"]:
            for name, value in zip(p["locations"], w["v"]):
                if value is None:
                    continue
                rows.append((w["start"], w["end"], name, f"{value:.6f}"))
                consolidated.append(
                    (p["label"], p["unit"], w["start"], w["end"], name, f"{value:.6f}")
                )
        with open(OUT_DIR / f"ppi_{key}.csv", "w", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerows(rows)
    with open(OUT_DIR / "ppi_long.csv", "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(consolidated)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Guardado antes de sobrescrever: as checagens de sanidade comparam a
    # execucao atual com a anterior para pegar serie que encolhe ou retrocede.
    anterior = None
    _prev = OUT_DIR / "ppi.json"
    if _prev.exists():
        try:
            with open(_prev, encoding="utf-8") as fh:
                anterior = json.load(fh)
        except Exception as exc:  # noqa: BLE001
            print(f"[qualidade] execucao anterior ilegivel, seguindo sem comparar: {exc}")
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Baixando {XLSX_URL}")
    blob = download(XLSX_URL)
    digest = hashlib.sha256(blob).hexdigest()
    print(f"  {len(blob):,} bytes | sha256 {digest[:16]}...")
    (RAW_DIR / "ppi.xlsx").write_bytes(blob)

    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    print(f"  abas encontradas: {wb.sheetnames}")

    products = {}
    for sheet_name, key, label in SHEETS:
        if sheet_name not in set(wb.sheetnames):
            token = norm(label.split()[0])
            candidates = [s for s in wb.sheetnames if token in norm(s)]
            if not candidates:
                raise SystemExit(f"aba '{sheet_name}' ausente na planilha da ANP")
            sheet_name = candidates[0]
        products[key] = parse_sheet(wb[sheet_name], key, label)
        p = products[key]
        print(
            f"  {label:<14} {len(p['weeks']):>4} semanas | "
            f"{len(p['locations']):>2} terminais | ate {p['weeks'][-1]['end']}"
        )

    latest = max(p["weeks"][-1]["end"] for p in products.values())
    dataset = {
        "generated_at": datetime.now(BR_TZ).isoformat(timespec="seconds"),
        "source_file": XLSX_URL,
        "source_page": PAGE_URL,
        "source_sha256": digest,
        "latest_week_end": latest,
        "order": [key for _, key, _ in SHEETS],
        "products": products,
    }

    with open(OUT_DIR / "ppi.json", "w", encoding="utf-8") as fh:
        json.dump(dataset, fh, ensure_ascii=False, separators=(",", ":"))

    with open(OUT_DIR / "meta.json", "w", encoding="utf-8") as fh:
        json.dump(
            {
                "generated_at": dataset["generated_at"],
                "latest_week_end": latest,
                "source_sha256": digest,
                "products": {
                    k: {
                        "weeks": len(p["weeks"]),
                        "locations": len(p["locations"]),
                        "unit": p["unit"],
                        "last_week": p["weeks"][-1]["label"],
                    }
                    for k, p in products.items()
                },
            },
            fh,
            ensure_ascii=False,
            indent=2,
        )

    write_csvs(products)

    try:
        import defasagem
        defasagem.run(products, OUT_DIR)
    except Exception as exc:  # noqa: BLE001
        print(f"[defasagem] etapa ignorada: {type(exc).__name__}: {exc}")

    try:
        import qualidade
        qualidade.run(products, anterior, OUT_DIR)
    except Exception as exc:  # noqa: BLE001
        print(f"[qualidade] etapa ignorada: {type(exc).__name__}: {exc}")

    size = (OUT_DIR / "ppi.json").stat().st_size
    print(f"OK -> docs/data/ppi.json ({size:,} bytes) | ultima semana {latest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
#!/usr/bin/env python3
"""
Baixa a planilha de Precos de Paridade de Importacao (PPI) da ANP,
normaliza as 4 abas semanais e gera os artefatos consumidos pelo dashboard.

Saidas (em docs/data/):
  ppi.json              -> dataset completo usado pelo dashboard
  ppi_<produto>.csv     -> serie longa por produto
  ppi_long.csv          -> serie longa consolidada dos 4 produtos
  meta.json             -> metadados da ultima execucao

Fonte: https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/
       precos/precos-de-paridade-de-importacao
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import sys
import unicodedata
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

XLSX_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/arq-ppi/ppi.xlsx"
)
PAGE_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/precos-de-paridade-de-importacao"
)

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "docs" / "data"
RAW_DIR = ROOT / "data_raw"

DATE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")
HEADER_LABELS = {"data", "semana"}

# Aba na planilha -> (chave interna, rotulo curto)
SHEETS = [
    ("Gasolina R$ semanal", "gasolina", "Gasolina A"),
    ("Diesel R$ semanal", "diesel", "Diesel A S10"),
    ("QAV R$ semanal", "qav", "QAV"),
    ("GLP R$ kg semanal", "glp", "GLP"),
]

BR_TZ = timezone(timedelta(hours=-3))


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #
def norm(text) -> str:
    """Minusculo, sem acentos e sem espacos duplicados."""
    if text is None:
        return ""
    s = str(text).replace("\xa0", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()


def clean(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).replace("\xa0", " ")).strip()


def as_number(value):
    """Converte celula em float, tolerando texto com virgula decimal e '-'."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return None if f != f else f
    s = str(value).strip()
    if s in {"", "-", "--", "n/d", "N/D", "ND"}:
        return None
    s = s.replace("R$", "").replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        f = float(s)
    except ValueError:
        return None
    return None if f != f else f


def download(url: str) -> bytes:
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; anp-ppi-dashboard/1.0)",
        "Accept": "*/*",
    }
    last = None
    for attempt in range(4):
        try:
            r = requests.get(url, headers=headers, timeout=120)
            r.raise_for_status()
            if len(r.content) < 10000:
                raise RuntimeError(f"arquivo truncado: {len(r.content)} bytes")
            return r.content
        except Exception as exc:  # noqa: BLE001
            last = exc
            print(f"  tentativa {attempt + 1} falhou: {exc}", file=sys.stderr)
    raise SystemExit(f"nao foi possivel baixar {url}: {last}")


# --------------------------------------------------------------------------- #
# parsing
# --------------------------------------------------------------------------- #
def find_header(rows):
    """Linha de cabecalho ('Data' ou 'Semana') e indice da coluna de data."""
    for i, row in enumerate(rows[:15]):
        for j, cell in enumerate(row):
            if norm(cell) in HEADER_LABELS:
                return i, j
    raise ValueError("cabecalho ('Data'/'Semana') nao encontrado")


def read_locations(header_row, date_col: int):
    """Terminais = celulas nao vazias apos a data, ate a primeira lacuna.

    A ANP repete os mesmos terminais num segundo bloco (variacao % semanal),
    separado por uma coluna em branco; paramos na lacuna.
    """
    locations = []
    for j in range(date_col + 1, len(header_row)):
        name = clean(header_row[j])
        if not name:
            break
        locations.append({"col": j, "name": name})
    if not locations:
        raise ValueError("nenhum terminal encontrado no cabecalho")
    return locations


def parse_period(raw):
    """'05/11/2018 a 09/11/2018' -> (inicio ISO, fim ISO, rotulo)."""
    if isinstance(raw, datetime):
        iso = raw.date().isoformat()
        return iso, iso, raw.strftime("%d/%m/%Y")
    label = clean(raw)
    matches = DATE_RE.findall(label)
    if not matches:
        return None

    def iso(m):
        d, mth, y = m
        return f"{y}-{mth}-{d}"

    start, end = iso(matches[0]), iso(matches[-1])
    try:
        datetime.fromisoformat(start)
        datetime.fromisoformat(end)
    except ValueError:
        return None
    return start, end, label


def parse_sheet(ws, key: str, fallback_label: str) -> dict:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header_idx, date_col = find_header(rows)
    locations = read_locations(rows[header_idx], date_col)

    def first_text(idx):
        if idx < 0 or idx >= len(rows):
            return ""
        for cell in rows[idx]:
            txt = clean(cell)
            if txt and norm(txt) != "variacao % em relacao a semana anterior":
                return txt
        return ""

    title = first_text(0) or fallback_label
    unit_raw = first_text(1)
    m = re.search(r"(R\$\s*/\s*[^\s,;]+)", unit_raw)
    unit = clean(m.group(1)) if m else (unit_raw or "R$")

    weeks, seen = [], set()
    for row in rows[header_idx + 1:]:
        if date_col >= len(row):
            continue
        period = parse_period(row[date_col])
        if not period:
            continue
        start, end, label = period
        if end in seen:
            continue
        values = []
        for loc in locations:
            v = as_number(row[loc["col"]]) if loc["col"] < len(row) else None
            values.append(round(v, 6) if v is not None else None)
        if all(v is None for v in values):
            continue
        seen.add(end)
        weeks.append({"start": start, "end": end, "label": label, "v": values})

    weeks.sort(key=lambda w: w["end"])
    if not weeks:
        raise ValueError(f"aba '{ws.title}' sem linhas de dados validas")

    return {
        "key": key,
        "label": fallback_label,
        "title": title,
        "unit": unit,
        "sheet": ws.title,
        "locations": [loc["name"] for loc in locations],
        "weeks": weeks,
    }


# --------------------------------------------------------------------------- #
# saidas
# --------------------------------------------------------------------------- #
def write_csvs(products: dict):
    consolidated = [
        ("produto", "unidade", "semana_inicio", "semana_fim", "terminal", "preco")
    ]
    for key, p in products.items():
        rows = [("semana_inicio", "semana_fim", "terminal", "preco")]
        for w in p["weeks"]:
            for name, value in zip(p["locations"], w["v"]):
                if value is None:
                    continue
                rows.append((w["start"], w["end"], name, f"{value:.6f}"))
                consolidated.append(
                    (p["label"], p["unit"], w["start"], w["end"], name, f"{value:.6f}")
                )
        with open(OUT_DIR / f"ppi_{key}.csv", "w", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerows(rows)
    with open(OUT_DIR / "ppi_long.csv", "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(consolidated)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Baixando {XLSX_URL}")
    blob = download(XLSX_URL)
    digest = hashlib.sha256(blob).hexdigest()
    print(f"  {len(blob):,} bytes | sha256 {digest[:16]}...")
    (RAW_DIR / "ppi.xlsx").write_bytes(blob)

    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    print(f"  abas encontradas: {wb.sheetnames}")

    products = {}
    for sheet_name, key, label in SHEETS:
        if sheet_name not in set(wb.sheetnames):
            token = norm(label.split()[0])
            candidates = [s for s in wb.sheetnames if token in norm(s)]
            if not candidates:
                raise SystemExit(f"aba '{sheet_name}' ausente na planilha da ANP")
            sheet_name = candidates[0]
        products[key] = parse_sheet(wb[sheet_name], key, label)
        p = products[key]
        print(
            f"  {label:<14} {len(p['weeks']):>4} semanas | "
            f"{len(p['locations']):>2} terminais | ate {p['weeks'][-1]['end']}"
        )

    latest = max(p["weeks"][-1]["end"] for p in products.values())
    dataset = {
        "generated_at": datetime.now(BR_TZ).isoformat(timespec="seconds"),
        "source_file": XLSX_URL,
        "source_page": PAGE_URL,
        "source_sha256": digest,
        "latest_week_end": latest,
        "order": [key for _, key, _ in SHEETS],
        "products": products,
    }

    with open(OUT_DIR / "ppi.json", "w", encoding="utf-8") as fh:
        json.dump(dataset, fh, ensure_ascii=False, separators=(",", ":"))

    with open(OUT_DIR / "meta.json", "w", encoding="utf-8") as fh:
        json.dump(
            {
                "generated_at": dataset["generated_at"],
                "latest_week_end": latest,
                "source_sha256": digest,
                "products": {
                    k: {
                        "weeks": len(p["weeks"]),
                        "locations": len(p["locations"]),
                        "unit": p["unit"],
                        "last_week": p["weeks"][-1]["label"],
                    }
                    for k, p in products.items()
                },
            },
            fh,
            ensure_ascii=False,
            indent=2,
        )

    write_csvs(products)

    try:
        import defasagem
        defasagem.run(products, OUT_DIR)
    except Exception as exc:  # noqa: BLE001
        print(f"[defasagem] etapa ignorada: {type(exc).__name__}: {exc}")

    size = (OUT_DIR / "ppi.json").stat().st_size
    print(f"OK -> docs/data/ppi.json ({size:,} bytes) | ultima semana {latest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
