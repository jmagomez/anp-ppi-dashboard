#!/usr/bin/env python3
"""
Defasagem PPI x preco de realizacao (produtores/importadores).

Fonte do preco interno: ANP, "Precos Medios Ponderados Semanais" de produtores
e importadores de derivados de petroleo e biodiesel (serie a partir de 2013).
https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/
precos-de-produtores-e-importadores-de-derivados-de-petroleo-e-biodiesel

Esta versao esta em MODO DESCOBERTA: baixa o arquivo, identifica o formato e
grava em docs/data/_debug_ppidp.json um retrato da estrutura (abas, dimensoes,
primeiras linhas). Nenhuma falha aqui pode derrubar o pipeline principal.
"""

from __future__ import annotations

import io
import json
import re
import unicodedata
from pathlib import Path

import requests

PPIDP_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/ppidp/precos-medios-ponderados-semanais-2013.xls"
)
PPIDP_PAGE = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/precos-de-produtores-e-importadores-de-derivados-de-petroleo-e-biodiesel"
)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; anp-ppi-dashboard/1.0)",
    "Accept": "*/*",
}


def norm(text) -> str:
    if text is None:
        return ""
    s = str(text).replace("\xa0", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()


def download(url: str) -> bytes:
    last = None
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=180)
            r.raise_for_status()
            if len(r.content) < 5000:
                raise RuntimeError(f"arquivo truncado: {len(r.content)} bytes")
            return r.content
        except Exception as exc:  # noqa: BLE001
            last = exc
            print(f"  [ppidp] tentativa {attempt + 1} falhou: {exc}")
    raise RuntimeError(f"download falhou: {last}")


def cell_repr(v):
    """Representacao curta e serializavel de uma celula."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace("\xa0", " ").strip()
    return s[:60] if s else None


def inspect(blob: bytes) -> dict:
    magic = blob[:8]
    info = {
        "bytes": len(blob),
        "magic": magic.hex(),
        "format": None,
        "sheets": [],
        "error": None,
    }

    # OLE2 (.xls antigo) -> xlrd ; ZIP (.xlsx) -> openpyxl
    if magic.startswith(b"\xd0\xcf\x11\xe0"):
        info["format"] = "xls (OLE2)"
        import xlrd

        wb = xlrd.open_workbook(file_contents=blob)
        for ws in wb.sheets():
            rows = []
            for r in range(min(ws.nrows, 14)):
                rows.append([cell_repr(ws.cell_value(r, c))
                             for c in range(min(ws.ncols, 14))])
            info["sheets"].append({
                "name": ws.name,
                "nrows": ws.nrows,
                "ncols": ws.ncols,
                "head": rows,
            })
    elif magic.startswith(b"PK"):
        info["format"] = "xlsx (ZIP)"
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 14:
                    break
                rows.append([cell_repr(v) for v in row[:14]])
            info["sheets"].append({
                "name": name,
                "nrows": ws.max_row,
                "ncols": ws.max_column,
                "head": rows,
            })
    else:
        info["format"] = "desconhecido"
        info["preview"] = blob[:400].decode("latin-1", "replace")

    return info


def run(products: dict, out_dir: Path) -> None:
    """Executado ao final do build principal. Nunca levanta excecao."""
    debug = {"source_file": PPIDP_URL, "source_page": PPIDP_PAGE}
    try:
        print(f"[ppidp] baixando {PPIDP_URL}")
        blob = download(PPIDP_URL)
        print(f"[ppidp] {len(blob):,} bytes")
        debug.update(inspect(blob))
        print(f"[ppidp] formato: {debug.get('format')} | "
              f"abas: {[s['name'] for s in debug.get('sheets', [])]}")
    except Exception as exc:  # noqa: BLE001
        debug["error"] = f"{type(exc).__name__}: {exc}"
        print(f"[ppidp] ERRO: {debug['error']}")

    with open(out_dir / "_debug_ppidp.json", "w", encoding="utf-8") as fh:
        json.dump(debug, fh, ensure_ascii=False, indent=1)
